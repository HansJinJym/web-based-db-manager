from openpyxl import load_workbook

# 当前登录用户
current_user = ''

# 记录表格，只读
op_record = '../record.xlsx'

def record_operation(data, xlsx_file=op_record):
    wb = load_workbook(filename=xlsx_file)
    ws = wb.active

    num = 1
    while 1:
        cell = ws.cell(row=num, column=1).value
        if cell:
            num = num +1
        else:
            row_num = num
            break

    for i in range(len(data)):
        ws.cell(row=row_num, column=i+1, value=data[i])

    wb.save(xlsx_file)
