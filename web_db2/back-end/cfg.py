from openpyxl import load_workbook

# {"code":"002594","name":"比亚迪","trade":"sell","price":"333.22","amount":"100"}
# {"code":"600030","name":"中信证券","trade":"buy","price":"20.22","amount":"10200"}
# {"cat":"股指期货","code":"IF2209","name":"沪深2209","trade":"买多","price":"4261.0","amount":"200"}
# {"username":"admin4","password":"1qaz2wsx"}

# 当前登录用户
current_user = ''

# 当前表格
db_path = '../stock.db'

# 记录表格，只读
op_record = '../record.xlsx'

def record_operation(data, xlsx_file=op_record):
    wb = load_workbook(filename=xlsx_file)
    ws = wb.active

    num = 1
    while 1:
        cell = ws.cell(row=num, column=1).value
        if cell:
            num = num +1
        else:
            row_num = num
            break

    for i in range(len(data)):
        ws.cell(row=row_num, column=i+1, value=data[i])

    wb.save(xlsx_file)
